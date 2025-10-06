#
#   Ejemplo 01: Crear un archivo Excel con Python, usando la librería openpyxl
#

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

# Crear un nuevo libro de Excel
wb = openpyxl.Workbook()
# Seleccionar la hoja activa
hoja = wb.active
hoja.title = "Ejemplo01"
# Cambiar el nombre de la hoja activa
# hoja = wb.create_sheet("Ejemplo01")  # Otra forma de crear una hoja
# hoja = wb["Ejemplo01"]  # Otra forma de seleccionar una hoja
# hoja = wb.worksheets[0]  # Otra forma de seleccionar una hoja
# Eliminar una hoja
# wb.remove(wb["Sheet"])  # Eliminar la hoja por nombre (la hoja por defecto)
# wb.remove(wb.worksheets[0])  # Eliminar la hoja por índice (la hoja por defecto)

# Escribir datos en celdas
hoja["A1"] = "Nombre"
hoja["B1"] = "Edad"
hoja["A2"] = "Juan"
hoja["B2"] = 30
hoja["A3"] = "María"
hoja["B3"] = 25
hoja["A4"] = "Pedro"
hoja["B4"] = 35
hoja["A5"] = "Ana"
hoja["B5"] = 28

# Aplicar estilos a las celdas
negrita = Font(bold=True)
centrado = Alignment(horizontal="center")
relleno_cabecera = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
hoja["A1"].font = negrita
hoja["B1"].font = negrita
hoja["A1"].alignment = centrado
hoja["B1"].alignment = centrado
hoja["A1"].fill = relleno_cabecera
hoja["B1"].fill = relleno_cabecera
# Ajustar el ancho de las columnas
hoja.column_dimensions["A"].width = 15
hoja.column_dimensions["B"].width = 10
# Ajustar la altura de las filas
hoja.row_dimensions[1].height = 20
hoja.row_dimensions[2].height = 18
hoja.row_dimensions[3].height = 18
hoja.row_dimensions[4].height = 18
hoja.row_dimensions[5].height = 18

# Alinear a la izquierda la columna B desde la fila 2 en adelante
alinear_izquierda = Alignment(horizontal="left")
alinear_derecha   = Alignment(horizontal="right")
for fila in range(2, 6):  # De la fila 2 a la 5
    hoja[f"B{fila}"] = str(hoja[f"B{fila}"].value)  # Convertir a texto
    hoja[f"B{fila}"].alignment = alinear_derecha
    #hoja[f"B{fila}"].alignment = alinear_izquierda

# Guardar el archivo Excel
wb.save("Ejemplo01.xlsx")
print("Archivo 'Ejemplo01.xlsx' creado con éxito.")
# Cerrar el libro de Excel
wb.close()
# print("Archivo 'Ejemplo01.xlsx' cerrado con éxito.")
# Nota: El archivo se guarda en el mismo directorio donde se encuentra este script.